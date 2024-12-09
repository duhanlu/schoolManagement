import openpyxl

# read excel 
class ReadExcel():
    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)
        self.worksheet = self.workbook.active
    
    def get_data(self):
        data = []
        for row in self.worksheet.iter_rows():
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            data.append(row_value)
        return data

class WriteExcel():
    def __init__(self, file_path, data):
        self.data = data
        self.file_path = file_path
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    
    def write_to_excel(self):
        for row in self.data:
            self.worksheet.append(row)
        self.workbook.save(self.file_path)


